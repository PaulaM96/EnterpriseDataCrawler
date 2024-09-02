import os
import logging
import tkinter as tk
from tkinter import messagebox, ttk
from seleniumbase import SB
from openpyxl import load_workbook, Workbook
import re
import threading
from collections import deque

# Configuração do log
LOG_FILE = "programa.log"
logging.basicConfig(level=logging.INFO, filename=LOG_FILE, format="%(asctime)s - %(levelname)s - %(message)s")

# Configurações gerais
CANCELAR_BUSCA = False
TOTAL_LINKS_ENCONTRADOS = 0
CAMINHO_CHROMEDRIVER = os.path.join(os.path.dirname(__file__), 'drivers', 'chromedriver.exe')


def inicializar_planilha(caminho_planilha, cabecalho):
    """Inicializa uma planilha no caminho especificado com o cabeçalho fornecido."""
    if not os.path.exists(os.path.dirname(caminho_planilha)):
        os.makedirs(os.path.dirname(caminho_planilha))

    if not os.path.exists(caminho_planilha):
        wb = Workbook()
        ws = wb.active
        ws.append(cabecalho)
        wb.save(caminho_planilha)
    return load_workbook(caminho_planilha)


def salvar_dados_excel(caminho_planilha, dados):
    """Salva os dados coletados em uma planilha Excel."""
    wb = load_workbook(caminho_planilha)
    ws = wb.active
    ws.append(dados)
    wb.save(caminho_planilha)
    logging.info(f"Dados salvos em: {caminho_planilha}")


def buscar_dados(sb, xpath, default="Não informado"):
    """Busca um elemento na página pelo XPath e retorna seu texto ou um valor padrão se não encontrado."""
    try:
        sb.wait_for_element_present(xpath, timeout=15)
        return sb.get_text(xpath)
    except:
        return default


def coletar_dados_empresa(sb, url, progress_var, total_urls, window):
    """Coleta os dados de uma empresa a partir de uma URL especificada."""
    global CANCELAR_BUSCA

    if CANCELAR_BUSCA:
        return None

    for tentativa in range(2):
        try:
            sb.open(url)
            sb.sleep(2)
            sb.wait_for_ready_state_complete()

            razao_social = buscar_dados(sb, '//p[contains(text(),"Razão Social")]/span/b')
            nome_fantasia = buscar_dados(sb, '//p[contains(text(),"Nome Fantasia")]/span/b', default=razao_social)

            # Quadro de Sócios e Administradores
            socios = buscar_dados(sb, '//h2[contains(text(),"Quadro de Sócios e Administradores")]/following-sibling::p')
            socios = ', '.join(socios.split('\n')).strip() if socios != "Não informado" else socios

            # Telefones
            telefones = coletar_telefones(sb)

            # E-mail
            email = buscar_dados(sb, '//p[contains(text(),"E-mail")]/span/b')

            # CNPJ
            cnpj = buscar_dados(sb, '//p[contains(text(),"CNPJ")]/span[2]/b')

            # Cidade e Estado
            cidade = buscar_dados(sb, '//p[contains(text(),"Município")]/b/a')
            estado = buscar_dados(sb, '//p[contains(text(),"Estado")]/b/a')
            cidade_uf = f"{cidade}/{estado}"

            # Retornar os dados coletados
            return [nome_fantasia, razao_social, socios, telefones, email, cnpj, cidade_uf]

        except Exception as e:
            logging.error(f"Erro ao coletar dados da empresa em {url} (Tentativa {tentativa + 1}): {e}")
        finally:
            progress_var.set(progress_var.get() + ((100 - 10) / total_urls))
            window.update_idletasks()

    return None


def coletar_telefones(sb):
    """Coleta números de telefone da página."""
    telefones = []
    elementos = sb.find_elements('//p[contains(text(),"Telefone")]/span/b')
    for elemento in elementos:
        numero = elemento.text
        if re.match(r'\(\d{2}\) \d{4,5}-\d{4}', numero):
            telefones.append(numero)
    return ", ".join(telefones)


def buscar_urls_empresas(sb, termo_pesquisa, window, progress_label):
    """Busca URLs de empresas ativas a partir de um termo de pesquisa."""
    global TOTAL_LINKS_ENCONTRADOS, CANCELAR_BUSCA
    sb.open("https://cnpj.biz/empresas")
    sb.wait_for_ready_state_complete()
    sb.type("#q", termo_pesquisa)
    sb.click("button[type='submit']")
    sb.sleep(2)
    sb.wait_for_ready_state_complete()

    urls_empresas = set()
    botao_proxima_pagina_xpath = "//a[contains(text(), 'Próxima Página')]"
    while not CANCELAR_BUSCA:
        try:
            empresas_ativas = sb.find_elements("li:has(p.bg-green-100.text-green-800) a")
            for empresa in empresas_ativas:
                url = empresa.get_attribute("href")
                if url and url not in urls_empresas:
                    urls_empresas.add(url)
                    TOTAL_LINKS_ENCONTRADOS += 1
                    progress_label.config(text=f"URLs encontradas: {TOTAL_LINKS_ENCONTRADOS}")
                    window.update_idletasks()

            if not navegar_para_proxima_pagina(sb, botao_proxima_pagina_xpath):
                break

        except Exception as e:
            logging.error(f"Erro ao tentar navegar para a próxima página: {e}")
            break

    return list(urls_empresas)


def navegar_para_proxima_pagina(sb, botao_xpath):
    """Navega para a próxima página se o botão estiver presente."""
    proxima_pagina_elemento = sb.find_elements(botao_xpath)
    if proxima_pagina_elemento:
        sb.scroll_to(botao_xpath)
        sb.click(botao_xpath)
        sb.wait_for_ready_state_complete()
        sb.sleep(1)
        return True
    logging.info("Nenhum botão de 'Próxima Página' encontrado, finalizando a busca.")
    return False


def verificar_e_deletar_planilha(caminho_planilha):
    """Verifica e deleta a planilha existente."""
    if os.path.exists(caminho_planilha):
        try:
            os.remove(caminho_planilha)
            logging.info(f"Arquivo {caminho_planilha} deletado com sucesso.")
        except Exception as e:
            logging.error(f"Erro ao tentar deletar o arquivo {caminho_planilha}: {e}")


def executar_busca(termo_pesquisa, progress_var, window, search_btn, abrir_pasta_btn, progress_label):
    """Executa a busca de empresas a partir do termo de pesquisa informado."""
    global CANCELAR_BUSCA, TOTAL_LINKS_ENCONTRADOS
    progress_var.set(0)
    CANCELAR_BUSCA = False
    TOTAL_LINKS_ENCONTRADOS = 0
    progress_label.config(text="URLs encontradas: 0")

    caminho_planilha = os.path.join(r"C:\DADOS_CNPJ", f"cnpjs_ativos_{termo_pesquisa}.xlsx")
    cabecalho = ["Nome Fantasia", "Razão Social", "Nome dos Sócios", "Telefones", "Email", "CNPJ", "Cidade/UF"]

    verificar_e_deletar_planilha(caminho_planilha)

    search_btn.config(state=tk.DISABLED)
    abrir_pasta_btn.config(state=tk.DISABLED)
    progress_var.set(10)
    window.update_idletasks()

    urls_empresas = deque()

    with SB(uc=True, headless=True) as sb:
        urls_empresas.extend(buscar_urls_empresas(sb, termo_pesquisa, window, progress_label))

    if urls_empresas:
        total_urls = len(urls_empresas)
        planilha = inicializar_planilha(caminho_planilha, cabecalho)

        with SB(uc=True, headless=True) as sb:
            while urls_empresas and not CANCELAR_BUSCA:
                url_atual = urls_empresas.popleft()
                dados_empresa = coletar_dados_empresa(sb, url_atual, progress_var, total_urls, window)
                if dados_empresa:
                    salvar_dados_excel(caminho_planilha, dados_empresa)

    finalizar_busca(progress_var, window, search_btn, abrir_pasta_btn, caminho_planilha)


def finalizar_busca(progress_var, window, search_btn, abrir_pasta_btn, caminho_planilha):
    """Finaliza a busca e atualiza a interface."""
    progress_var.set(100)
    window.update_idletasks()
    messagebox.showinfo("Concluído", f"Busca concluída. Dados salvos em {caminho_planilha}")
    abrir_pasta_btn.config(state=tk.NORMAL, command=lambda: abrir_pasta(caminho_planilha))
    search_btn.config(state=tk.NORMAL)


def cancelar_busca_func():
    """Cancela a busca de empresas."""
    global CANCELAR_BUSCA
    CANCELAR_BUSCA = True


def abrir_pasta(caminho_planilha):
    """Abre o diretório onde a planilha está salva."""
    os.startfile(os.path.dirname(caminho_planilha))


def iniciar_interface():
    """Inicia a interface gráfica do usuário."""
    window = tk.Tk()
    window.title("Consulta de Empresas")
    window.geometry("400x400")

    termo_label = tk.Label(window, text="Digite o termo de pesquisa:")
    termo_label.pack(pady=10)

    termo_entry = tk.Entry(window)
    termo_entry.pack(pady=10)

    progress_var = tk.DoubleVar()
    progress_bar = ttk.Progressbar(window, variable=progress_var, maximum=100)
    progress_bar.pack(pady=20, fill=tk.X)

    progress_label = tk.Label(window, text="URLs encontradas: 0")
    progress_label.pack(pady=10)

    abrir_pasta_btn = tk.Button(window, text="Abrir Pasta", state=tk.DISABLED)
    abrir_pasta_btn.pack(pady=10)

    search_btn = tk.Button(
        window,
        text="Buscar",
        command=lambda: threading.Thread(
            target=executar_busca,
            args=(termo_entry.get(), progress_var, window, search_btn, abrir_pasta_btn, progress_label)
        ).start()
    )
    search_btn.pack(pady=10)

    cancelar_btn = tk.Button(window, text="Cancelar", command=cancelar_busca_func)
    cancelar_btn.pack(pady=10)

    window.mainloop()


if __name__ == "__main__":
    iniciar_interface()
